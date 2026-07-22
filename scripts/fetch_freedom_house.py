import json
import io
import sys
from pathlib import Path
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).resolve().parent))
from fetch_common import (
    FETCHED,
    ROOT,
    USER_AGENT,
    ensure_fetched_dir,
    exit_for_status,
    fetch_bytes,
    write_status,
)

ensure_fetched_dir()

CACHE_XLSX = FETCHED / "freedom_house.xlsx"
OUTPUT_JSON = FETCHED / "freedom_house.json"

URLS = [
    "https://freedomhouse.org/sites/default/files/2025-02/All_data_FIW_2013-2024.xlsx",
    "https://freedomhouse.org/sites/default/files/2024-03/All_data_FIW_2013-2023.xlsx",
]


def download_xlsx(url: str) -> bytes | None:
    data, err = fetch_bytes(url, user_agent=USER_AGENT)
    if err:
        print(f"  failed: {url} -- {err}")
        return None
    return data


def parse_with_openpyxl(raw: bytes) -> dict | None:
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed; cannot parse xlsx")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        print(f"  openpyxl failed to open workbook: {e}")
        return None

    countries = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        header_idx = None
        for i, row in enumerate(rows):
            vals = [str(c).strip().lower() if c else "" for c in row]
            if "country/territory" in vals or "country" in vals:
                header_idx = i
                break
        if header_idx is None:
            continue

        header = [str(c).strip().lower() if c else "" for c in rows[header_idx]]

        def find_col(*candidates):
            for cand in candidates:
                for i, h in enumerate(header):
                    if h == cand:
                        return i
            for cand in candidates:
                for i, h in enumerate(header):
                    if cand in h:
                        return i
            return None

        col_country = find_col("country/territory", "country", "nation")
        col_status = find_col("status")
        col_pr = find_col("pr rating", "pr")
        col_cl = find_col("cl rating", "cl")
        col_year = find_col("edition", "year")

        if col_country is None or col_status is None:
            continue

        STATUS_MAP = {
            "f": "Free", "free": "Free",
            "pf": "Partly Free", "partly free": "Partly Free",
            "nf": "Not Free", "not free": "Not Free",
        }

        for row in rows[header_idx + 1:]:
            vals = list(row)
            if len(vals) <= max(col_country, col_status):
                continue
            country_name = vals[col_country]
            if not country_name or not isinstance(country_name, str):
                continue
            country_name = country_name.strip()
            if not country_name or country_name.lower() in ("total", "avg", "average", ""):
                continue

            def safe_int(idx):
                if idx is None or idx >= len(vals):
                    return None
                v = vals[idx]
                if v is None:
                    return None
                try:
                    return int(v)
                except (ValueError, TypeError):
                    return None

            def safe_str(idx):
                if idx is None or idx >= len(vals):
                    return None
                v = vals[idx]
                if v is None:
                    return None
                return str(v).strip()

            raw_status = safe_str(col_status)
            status = STATUS_MAP.get(raw_status.lower(), raw_status if raw_status else None) if raw_status else None
            pr = safe_int(col_pr)
            cl = safe_int(col_cl)
            year = safe_int(col_year)

            if country_name not in countries:
                countries[country_name] = {
                    "status": status,
                    "pr_score": pr,
                    "cl_score": cl,
                    "year": year,
                }
            else:
                existing = countries[country_name]
                if year and (existing["year"] is None or year > existing["year"]):
                    countries[country_name] = {
                        "status": status,
                        "pr_score": pr,
                        "cl_score": cl,
                        "year": year,
                    }

    wb.close()
    return countries if countries else None


def parse_with_csv_fallback(raw: bytes) -> dict | None:
    import csv
    try:
        text = raw.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if len(rows) < 2:
            return None

        header = [h.strip().lower() for h in rows[0]]

        def find_col(*candidates):
            for cand in candidates:
                for i, h in enumerate(header):
                    if cand in h:
                        return i
            return None

        col_country = find_col("country", "nation")
        col_status = find_col("status", "freedom status")
        col_pr = find_col("pr", "political rights")
        col_cl = find_col("cl", "civil liberties")
        col_year = find_col("year", "edition")

        if col_country is None:
            return None

        countries = {}
        for row in rows[1:]:
            if len(row) <= col_country:
                continue
            country_name = row[col_country].strip()
            if not country_name or country_name.lower() in ("total", "avg", "average"):
                continue

            def safe_int(idx):
                if idx is None or idx >= len(row):
                    return None
                v = row[idx].strip()
                if not v:
                    return None
                try:
                    return int(v)
                except ValueError:
                    return None

            def safe_str(idx):
                if idx is None or idx >= len(row):
                    return None
                return row[idx].strip() or None

            status = safe_str(col_status)
            pr = safe_int(col_pr)
            cl = safe_int(col_cl)
            year = safe_int(col_year)

            if status and "not free" in status.lower():
                status = "Not Free"
            elif status and "partly free" in status.lower():
                status = "Partly Free"
            elif status and "free" in status.lower():
                status = "Free"

            if country_name not in countries:
                countries[country_name] = {
                    "status": status,
                    "pr_score": pr,
                    "cl_score": cl,
                    "year": year,
                }
            else:
                existing = countries[country_name]
                if year and (existing["year"] is None or year > existing["year"]):
                    countries[country_name] = {
                        "status": status,
                        "pr_score": pr,
                        "cl_score": cl,
                        "year": year,
                    }

        return countries if countries else None
    except Exception as e:
        print(f"  csv fallback failed: {e}")
        return None


def main():
    print("Freedom House Freedom in the World data fetcher")
    print("=" * 50)

    raw = None
    used_url = None
    for url in URLS:
        print(f"\nTrying: {url}")
        data = download_xlsx(url)
        if data:
            raw = data
            used_url = url
            CACHE_XLSX.write_bytes(raw)
            print(f"  downloaded ({len(raw):,} bytes), cached to {CACHE_XLSX.relative_to(ROOT)}")
            break

    if raw is None and CACHE_XLSX.exists():
        print(f"\nAll downloads failed, using cached: {CACHE_XLSX.relative_to(ROOT)}")
        raw = CACHE_XLSX.read_bytes()
        fetch_status = "cached"
    elif raw is None:
        print("\nNo data available and no cache found.")
        result = {"countries": {}, "fetched_at": None, "error": "no data available"}
        OUTPUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
        write_status("freedomhouse", "failed", "no data available")
        exit_for_status("failed")
    else:
        fetch_status = "ok"

    print("\nParsing with openpyxl...")
    countries = parse_with_openpyxl(raw)

    if countries is None:
        print("  openpyxl parse failed, trying csv fallback...")
        countries = parse_with_csv_fallback(raw)

    if countries is None:
        print("  all parsers failed.")
        result = {"countries": {}, "fetched_at": None, "error": "parse failed"}
        OUTPUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
        write_status("freedomhouse", "failed", "parse failed")
        exit_for_status("failed")

    result = {
        "countries": countries,
        "fetched_at": datetime.now(timezone.utc).isoformat(),
        "source_url": used_url,
        "total_countries": len(countries),
    }
    OUTPUT_JSON.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    write_status("freedomhouse", fetch_status)

    print(f"\n{'=' * 50}")
    print(f"Total countries: {len(countries)}")

    status_counts = {}
    for c in countries.values():
        s = c.get("status") or "Unknown"
        status_counts[s] = status_counts.get(s, 0) + 1
    for s, n in sorted(status_counts.items()):
        print(f"  {s}: {n}")

    years = [c["year"] for c in countries.values() if c.get("year")]
    if years:
        print(f"Year range: {min(years)}-{max(years)}")

    print(f"\nOutput: {OUTPUT_JSON.relative_to(ROOT)}")
    print("\nDone.")
    exit_for_status(fetch_status)


if __name__ == "__main__":
    main()
